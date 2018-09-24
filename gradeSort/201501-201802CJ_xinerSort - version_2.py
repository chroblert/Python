#!/usr/bin/python
# _*_coding: utf-8 _*_
"""
created by Jerrybird

"""
import re
import chardet
import json
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf-8')

def htmlToJson(in_file_name,out_file_name='CJ.json'):
    with open(in_file_name,'r') as f:
        line1=f.read()
    type=chardet.detect(line1)
    data=line1.decode('GB2312',errors='ignore')
    #matchObj = re.finditer(r'</style>.*?</div>(.*?)<br />.*?/>(.*?)<br />.*?class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?<style>', data, re.M | re.I|re.S)
    #matchObj = re.finditer(r'class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>', data, re.M | re.I|re.S)

    matchObj = re.finditer(r'</style>.*?95%;".*?</div>.*?:(.*?)<br />.*?/>.*?:(.*?)<br />(.*?)<style>', data, re.M | re.I|re.S)
    CJ_list=[]
    if matchObj:
        for m in matchObj:
            stu_dict = {}
            # print m.group(1),
            # print m.group(2),
            stu_dict.setdefault('学号',m.group(1).strip())
            stu_dict.setdefault('姓名', m.group(2).strip())
            cj_data=m.group(3)
            grade_list = []
            for i in re.findall(r'class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>', cj_data, re.M | re.I|re.S):
                temp_list=[]
                for j in i:
                    #过滤
                    if '<strong>#</strong>' in j:
                        j=j.replace('<strong>#</strong>','#')
                    if '<strong>*</strong>' in j:
                        j=j.replace('<strong>*</strong>','*')
                    temp_list.append(j.strip())
                    # print j,
                # print
                grade_list.append(temp_list)
                stu_dict.setdefault('grade_list', grade_list)
            CJ_list.append(stu_dict)
            # print
    else:
        print "No match"

    #导出到json文件中
    with open(out_file_name,'w') as f:
        json.dump(CJ_list,f,ensure_ascii=False,indent=4)

def convertToTitle(n):
    """
    :type n: int
    :rtype: str
    #需要注意26时：26%26为0 也就是0为A 所以使用n-1  A的ASCII码为65
    """
    result = ""
    while n != 0:
        result = chr((n - 1) % 26 + 65) + result
        n = (n - 1) / 26
    return result

def jsonToExcel(in_file_name='CJ.json',out_file_name='CJ.xlsx'):
    #读取json文件
    with open(in_file_name,'r') as f:
        CJ_list=json.load(f)
    #导出到excel表
    #创建Excel
    wb=Workbook()
    #创建表
    sheet=wb.active
    sheet.title='CJ'

    row_i=1
    for s_dict in CJ_list:#每一个人的信息
        # kc_mc=''
        kc_list = list()
        col_i = 1
        xf=0
        cfkccs_dict = dict()
        for kc in s_dict['grade_list']:#每一个人的课程成绩单
            #过滤掉*和#，去掉两头空格
            bjggkc_list=list()
            if '#' in kc[0]:
                kc_mc=kc[0].replace('#','').strip()
                bjggkc_list.append(kc_mc)#不及格过的课程
            elif '*' in kc[0]:
                kc_mc=kc[0].replace('*','').strip()
                bjggkc_list.append(kc_mc)
            else:
                kc_mc = kc[0].strip()
                if kc_mc not in bjggkc_list and kc_mc in kc_list:  # 该课程没有不及格过且之前已经录入一次，则该课程要么是刷绩点，要么是一样名称的体育
                    # print kc_mc
                    temp=kc_mc
                    kc_mc = kc_mc + str(cfkccs_dict[kc_mc])
                    cfkccs_dict[temp] = cfkccs_dict[temp] + 1
                    # print

            if kc_mc not in kc_list:
                # 第一行课程名称
                sheet[convertToTitle(col_i +2) + str(row_i)] = kc_mc
                #第二行学分
                sheet[convertToTitle(col_i +2) + str(row_i + 1)] = kc[1]
                #第三行成绩
                sheet[convertToTitle(col_i +2 ) + str(row_i+2)] = kc[2]
                col_i = col_i + 1
                kc_list.append(kc_mc)
                cfkccs_dict.setdefault(kc_mc,1)
                xf = xf + float(kc[1])
            else:
                # print '####',kc_mc,kc[2],'####'
                if float(kc[2]) > float(sheet[convertToTitle(kc_list.index(kc_mc)+3) + str(row_i + 2)].value):
                    sheet[convertToTitle(kc_list.index(kc_mc)+3 ) + str(row_i +2 )].value = kc[2]
        del kc_list[:]
        # print s_dict[u'姓名'],xf
        sheet['A'+str(row_i)]=s_dict[u'学号']
        sheet['B' + str(row_i)] = s_dict[u'姓名']
        row_i = row_i + 3

    wb.save(out_file_name)

def excelToExcel(in_file_name='CJ.xlsx',out_file_name='CJ.xlsx'):
    wb=load_workbook(in_file_name)
    sheet=wb.active
    sheet_z=wb.create_sheet('sheet_z')
    sheet_z['A' + str(1)].value = '学号'
    sheet_z['B' + str(1)].value = '姓名'
    sheet_z['C' + str(1)].value = '总通过学分'
    sheet_z['D' + str(1)].value = '未通过学分'
    sheet_z['E' + str(1)].value = '总学分'
    sheet_z['F' + str(1)].value = '平均学分绩点（含公共选修课）'
    sheet_z['G' + str(1)].value = '平均学分绩点（不含公共选修课）'
    sheet_z['H' + str(1)].value = '未通过课程'
    i=1
    while (sheet['A' + str(i)].value != None):
        i = i + 3
    row_end=i-1#从3到end行都有数据
    for j in range(1,row_end+1,3):#每一个同学的,,行
        sum_tgxf=0
        sum_wtgxf = 0
        sum_xf=0
        sum_xfjd=0
        wtgkc_list=list()
        wtgkc_i=8
        k=3
        print sheet['B'+str(j)].value,
        while(sheet[convertToTitle(k)+str(j)].value != None):
            k=k+1
        col_end=k-1#从3到end列都有数据
        for i in range(3,col_end+1):#每一个成绩print sheet[convertToTitle(i)+str(j)].value
            # if sheet[convertToTitle(i)+str(j)].value != None:
            sum_xf=sum_xf + float(sheet[convertToTitle(i)+str(j+1)].value)
            if sheet[convertToTitle(i)+str(j+2)].value != None and float(sheet[convertToTitle(i)+str(j+2)].value) >= 60 :
                sum_tgxf = sum_tgxf + float(sheet[convertToTitle(i)+str(j+1)].value)
                sum_xfjd=sum_xfjd + ((float(sheet[convertToTitle(i)+str(j+2)].value)-50)/10) * float(sheet[convertToTitle(i)+str(j+1)].value)
                print str(((float(sheet[convertToTitle(i)+str(j+2)].value)-50)/10) * float(sheet[convertToTitle(i)+str(j+1)].value)),
            elif sheet[convertToTitle(i)+str(j+2)].value != None and float(sheet[convertToTitle(i)+str(j+2)].value) < 60:
                sum_wtgxf = sum_wtgxf + float(sheet[convertToTitle(i) + str(j+1)].value)
                # sheet_z[convertToTitle(wtgkc_i)+str(j-1)].value = sheet[convertToTitle(i)+str(j)].value
                wtgkc_list.append(sheet[convertToTitle(i)+str(j)].value+'('+sheet[convertToTitle(i)+str(j+1)].value+')')
                wtgkc_i = wtgkc_i + 1
        xfjd=sum_xfjd/sum_xf
        # print sheet['B'+str(j)].value,xfjd
        # print sum_xf,sum_tgxf,sum_wtgxf
        print

        sheet_z['A' + str(j//3 + 2)].value = sheet['A' + str(j)].value
        sheet_z['B' + str(j//3 + 2)].value = sheet['B' + str(j)].value
        sheet_z['C' + str(j//3 + 2)].value = str(sum_tgxf)
        sheet_z['D' + str(j//3 + 2)].value = str(sum_wtgxf)
        sheet_z['E' + str(j//3 + 2)].value = str(sum_xf)
        sheet_z['F' + str(j//3 + 2)].value = str(xfjd)
        sheet_z['G' + str(j//3 + 2)].value = ''
        # sheet_z[convertToTitle(wtgkc_i) + str(j - 1)].value = sheet[convertToTitle(i) + str(j)].value
        for n in wtgkc_list:
            if sheet_z[convertToTitle(9) + str(j//3 + 2)].value == None:
                sheet_z[convertToTitle(9) + str(j // 3 + 2)].value=str(wtgkc_list.index(n)+1) +'.'+ n +'\n'
            else:
                sheet_z[convertToTitle(9) + str(j//3 + 2)].value=  sheet_z[convertToTitle(9) + str(j//3 + 2)].value + str(wtgkc_list.index(n)+1)+'.' + n +'\n'
    wb.save(out_file_name)

def main():
    htmlToJson('CJ_sort.html')
    jsonToExcel()
    excelToExcel()

main()