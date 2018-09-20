#!/usr/bin/python
# _*_coding: utf-8 _*_

import re
import chardet
import json
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf-8')

def htmlToJson(in_file_name,out_file_name='CJ.json'):
    with open(in_file_name,'r') as f:
        line1=f.read()
    type=chardet.detect(line1)
    data=line1.decode('GB2312',errors='ignore')
    #matchObj = re.finditer(r'</style>.*?</div>(.*?)<br />.*?/>(.*?)<br />.*?class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?<style>', data, re.M | re.I|re.S)
    #matchObj = re.finditer(r'class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>', data, re.M | re.I|re.S)

    matchObj = re.finditer(r'</style>.*?95%;".*?</div>.*?:(.*?)<br />.*?/>.*?:(.*?)<br />(.*?)<style>', data, re.M | re.I|re.S)
    CJ_list=[]
    if matchObj:
        for m in matchObj:
            stu_dict = {}
            # print m.group(1),
            # print m.group(2),
            stu_dict.setdefault('学号',m.group(1).strip())
            stu_dict.setdefault('姓名', m.group(2).strip())
            cj_data=m.group(3)
            grade_list = []
            for i in re.findall(r'class="tableStyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>.*?StyleTd">(.*?)</td>', cj_data, re.M | re.I|re.S):
                temp_list=[]
                for j in i:
                    #过滤
                    if '<strong>#</strong>' in j:
                        j=j.replace('<strong>#</strong>','#')
                    if '<strong>*</strong>' in j:
                        j=j.replace('<strong>*</strong>','*')
                    temp_list.append(j.strip())
                    # print j,
                # print
                grade_list.append(temp_list)
                stu_dict.setdefault('grade_list', grade_list)
            CJ_list.append(stu_dict)
            # print
    else:
        print "No match"

    #导出到json文件中
    with open(out_file_name,'w') as f:
        json.dump(CJ_list,f,ensure_ascii=False,indent=4)

def convertToTitle(n):
    """
    :type n: int
    :rtype: str
    #需要注意26时：26%26为0 也就是0为A 所以使用n-1  A的ASCII码为65
    """
    result = ""
    while n != 0:
        result = chr((n - 1) % 26 + 65) + result
        n = (n - 1) / 26
    return result

def jsonToExcel(in_file_name='CJ.json',out_file_name='CJ.xlsx'):
    #读取json文件
    with open(in_file_name,'r') as f:
        CJ_list=json.load(f)
    #导出到excel表
    #创建Excel
    wb=Workbook()
    #创建表
    sheet=wb.active
    sheet.title='CJ'

    sheet['B1'].value='学分'
    sheet['A2'].value='学号'
    sheet['B2'].value='姓名'
    row_i=3
    kccol_i=3
    kc_list=list()
    for s_dict in CJ_list:
        sheet['A'+str(row_i)].value=s_dict[u'学号']
        sheet['B'+str(row_i)].value=s_dict[u'姓名']
        kc_mc=''
        for kc in s_dict['grade_list']:
            #过滤掉*和#，去掉两头空格
            if '#' in kc[0]:
                kc_mc=kc[0].replace('#','').strip()
            elif '*' in kc[0]:
                kc_mc=kc[0].replace('*','').strip()
            else:
                kc_mc=kc[0]
            print kc_mc
            if kc_mc not in kc_list:
                sheet[convertToTitle(kccol_i) + str(2)].value = kc_mc#录入课程名称
                sheet[convertToTitle(kccol_i) + str(1)].value = kc[1]#录入学分
                kccol_i = kccol_i + 1
                kc_list.append(kc_mc)
            col_i=kc_list.index(kc_mc)+3
            #读取单元格数字，若为空白则为0
            if sheet[convertToTitle(col_i) + str(row_i)].value == None or kc[2] > float(sheet[convertToTitle(col_i) + str(row_i)].value) :
                sheet[convertToTitle(col_i) + str(row_i)].value=str(kc[2])
        row_i=row_i+1
    wb.save(out_file_name)

def excelToExcel(in_file_name='CJ.xlsx',out_file_name='CJ.xlsx'):
    wb=load_workbook(in_file_name)
    sheet=wb.active
    sheet_z=wb.create_sheet('sheet_z')
    sheet_z['A' + str(1)].value = '学号'
    sheet_z['B' + str(1)].value = '姓名'
    sheet_z['C' + str(1)].value = '总通过学分'
    sheet_z['D' + str(1)].value = '未通过学分'
    sheet_z['E' + str(1)].value = '总学分'
    sheet_z['F' + str(1)].value = '平均学分绩点（含公共选修课）'
    sheet_z['G' + str(1)].value = '平均学分绩点（不含公共选修课）'
    sheet_z['H' + str(1)].value = '未通过课程'
    i=3
    while(sheet[convertToTitle(i)+str(2)].value != None):
        i=i+1
    col_end=i-1#从3到end都有数据
    i=3
    while (sheet['B' + str(i)].value != None):
        i = i + 1
    row_end=i-1
    for j in range(3,row_end+1):#每一个同学的
        sum_tgxf=0
        sum_wtgxf = 0
        sum_xf=0
        sum_xfjd=0
        wtgkc_list=list()
        wtgkc_i=8
        for i in range(3,col_end+1):#每一个成绩print sheet[convertToTitle(i)+str(j)].value
            if sheet[convertToTitle(i)+str(j)].value != None:
                sum_xf=sum_xf + float(sheet[convertToTitle(i)+str(1)].value)
            if sheet[convertToTitle(i)+str(j)].value != None and float(sheet[convertToTitle(i)+str(j)].value) >=60:
                sum_tgxf = sum_tgxf + float(sheet[convertToTitle(i)+str(1)].value)
                sum_xfjd=sum_xfjd + (float(sheet[convertToTitle(i)+str(j)].value)-50)/10 * float(sheet[convertToTitle(i)+str(1)].value)
            elif sheet[convertToTitle(i)+str(j)].value != None and float(sheet[convertToTitle(i)+str(j)].value) <=60:
                sum_wtgxf = sum_wtgxf + float(sheet[convertToTitle(i) + str(1)].value)
                sheet_z[convertToTitle(wtgkc_i)+str(j-1)].value = sheet[convertToTitle(i)+str(2)].value
                wtgkc_list.append(sheet[convertToTitle(i)+str(2)].value)
                wtgkc_i = wtgkc_i + 1
        xfjd=sum_xfjd/sum_xf
        print sheet['B'+str(j)].value,xfjd
        print sum_xf,sum_tgxf,sum_wtgxf
        sheet_z['A'+str(j-1)].value=sheet['A'+str(j)].value
        sheet_z['B' + str(j - 1)].value = sheet['B' + str(j)].value
        sheet_z['C'+str(j-1)].value=str(sum_tgxf)
        sheet_z['D' + str(j - 1)].value = str(sum_wtgxf)
        sheet_z['E' + str(j - 1)].value = str(sum_xf)
        sheet_z['F' + str(j - 1)].value = str(xfjd)
        sheet_z['G' + str(j - 1)].value = ''
    wb.save(out_file_name)

def main():
    htmlToJson('CJ_sort.html')
    jsonToExcel()
    excelToExcel()

main()