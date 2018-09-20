# !/usr/bin/python
# _*_coding: utf-8 _*_
import json
import sys
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf-8')
def convertToTitle(n):
    """
    :type n: int
    :rtype: str
    #需要注意26时：26%26为0 也就是0为A 所以使用n-1  A的ASCII码为65
    """
    result = ""
    while n != 0:
        result = chr((n - 1) % 26 + 65) + result
        n = (n - 1) / 26
    return result

def last():
    wb=load_workbook('cj.xlsx')
    sheet=wb.active
    sheet_z=wb.create_sheet('sheet_z')
    sheet_z['A' + str(1)].value = '学号'
    sheet_z['B' + str(1)].value = '姓名'
    sheet_z['C' + str(1)].value = '总通过学分'
    sheet_z['D' + str(1)].value = '未通过学分'
    sheet_z['E' + str(1)].value = '总学分'
    sheet_z['F' + str(1)].value = '平均学分绩点（含公共选修课）'
    sheet_z['G' + str(1)].value = '平均学分绩点（不含公共选修课）'
    sheet_z['H' + str(1)].value = '未通过课程'
    i=3
    while(sheet[convertToTitle(i)+str(2)].value != None):
        i=i+1
    col_end=i-1#从3到end都有数据
    i=3
    while (sheet['B' + str(i)].value != None):
        i = i + 1
    row_end=i-1
    for j in range(3,row_end+1):#每一个同学的
        sum_tgxf=0
        sum_wtgxf = 0
        sum_xf=0
        sum_xfjd=0
        wtgkc_list=list()
        wtgkc_i=8
        for i in range(3,col_end+1):#每一个成绩print sheet[convertToTitle(i)+str(j)].value
            if sheet[convertToTitle(i)+str(j)].value != None:
                sum_xf=sum_xf + float(sheet[convertToTitle(i)+str(1)].value)
            if sheet[convertToTitle(i)+str(j)].value != None and float(sheet[convertToTitle(i)+str(j)].value) >=60:
                sum_tgxf = sum_tgxf + float(sheet[convertToTitle(i)+str(1)].value)
                sum_xfjd=sum_xfjd + (float(sheet[convertToTitle(i)+str(j)].value)-50)/10 * float(sheet[convertToTitle(i)+str(1)].value)
            elif sheet[convertToTitle(i)+str(j)].value != None and float(sheet[convertToTitle(i)+str(j)].value) <=60:
                sum_wtgxf = sum_wtgxf + float(sheet[convertToTitle(i) + str(1)].value)
                sheet_z[convertToTitle(wtgkc_i)+str(j-1)].value = sheet[convertToTitle(i)+str(2)].value
                wtgkc_list.append(sheet[convertToTitle(i)+str(2)].value)
                wtgkc_i = wtgkc_i + 1
        xfjd=sum_xfjd/sum_xf
        print sheet['B'+str(j)].value,xfjd
        print sum_xf,sum_tgxf,sum_wtgxf
        sheet_z['A'+str(j-1)].value=sheet['A'+str(j)].value
        sheet_z['B' + str(j - 1)].value = sheet['B' + str(j)].value
        sheet_z['C'+str(j-1)].value=str(sum_tgxf)
        sheet_z['D' + str(j - 1)].value = str(sum_wtgxf)
        sheet_z['E' + str(j - 1)].value = str(sum_xf)
        sheet_z['F' + str(j - 1)].value = str(xfjd)
        sheet_z['G' + str(j - 1)].value = ''
    wb.save('cj.xlsx')
last()